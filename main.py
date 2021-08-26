import math
import openpyxl as op
from openpyxl.workbook import Workbook
from openpyxl.styles import Font

# open new workbook
# wb = op.Workbook()

# load workbook
# wb= op.load_workbook('worker tracing.xls')




categories = ["Date", "Hours", "Base salary", "Tip", "Overall", "Completion", "Additional hours",  "Shabbat hours", "comments" ]

worker_list = ["MichalD", "MichalH", "Yoad", "Alon", "Rinat", "Aime", "Nitai", "Gilad", "Noam", "Omri"]

months = ["january", "February", "march", "April", "May", "june", "July", "August", "September", "October", "Novembre",
          "December"]

income_list = ['date', 'x', "credit", "cash", "tip", "overall income"]
income_info = []




def add_sheet(wb, name, categories):
    ws = wb.create_sheet(name)
    for column, category in enumerate(categories):
        ws.cell(1, column + 1, category)



def find_first_empty(ws, date=''):
    current_row = 1
    while ws.cell(current_row, 1).value and ws.cell(current_row, 1).value != date:
        current_row += 1
    return current_row


def calculate_base_salary(hours, isPik, shabbat_hours):
    additional_hours = max(hours-8, 0)
    if shabbat_hours:
        additional_hours = 0
    base_sum = 35 if isPik else 30
    shabbat_sum = 45
    payment_for_regular_hours = (hours - additional_hours - shabbat_hours) * base_sum + (shabbat_hours * shabbat_sum)
    payment_for_additional_hours = additional_hours * (base_sum * 1.25)
    base_salary = payment_for_regular_hours + payment_for_additional_hours
    return math.ceil(base_salary)


def calculate_completion(base_salary, tip):
    completion = 100 if base_salary < tip + 100 else base_salary - tip
    return completion

def update_sum(ws, columns_count):
    current_row = find_first_empty(ws)
    sums = [0] * columns_count
    for i in range(2, current_row):
        for j in range(2, columns_count + 1):
            sums[j-1] += ws.cell(i, j).value
    sums[0] = 'sums'
    for i, sum in enumerate(sums):
        ws.cell(current_row + 3, i+ 1, '').font = Font(bold=False)
        ws.cell(current_row + 4, i+ 1, sum).font = Font(bold=True)



def update_worker(ws, date, hours, tip, shabbat_hours, isPik):
    additional_hours = max(hours-8, 0)
    base_salary = calculate_base_salary(hours, isPik, shabbat_hours)
    completion = calculate_completion(base_salary, tip)
    overall = tip + completion
    worker_details = [date, hours, base_salary, tip,  overall, completion, additional_hours, shabbat_hours, "pik" if isPik else ""]
    empty_row = find_first_empty(ws, date)
    for column, detail in enumerate(worker_details):
        ws.cell(empty_row, column + 1, detail)
    update_sum(ws, 8)


def update_income(ws, date, x, credit, tip):
    overall_income = x + tip
    cash = x - credit
    income_details = [date, x, credit, cash,  tip, overall_income]
    current_row = find_first_empty(ws, date)
    for column, detail in enumerate(income_details):
        ws.cell(current_row, column + 1, detail)
    update_sum(ws, 6)


def new_workbook(workers_list, month, year):
    wb = op.Workbook()
    for worker in workers_list:
        add_sheet(wb, worker, categories)
    add_sheet(wb, months[month - 1] + " 20" + str(year), income_list)
    return wb

def update_pik(date):
    pick = input('Please enter pik worker name, hours (shabbat hours) :')
    pick_details = pick.split()
    shabbar_hours = float(pick_details[2]) if len(pick_details) == 3 else 0
    update_worker(wb[pick_details[0]], date, float(pick_details[1]), 0, shabbar_hours, True)

def regular_shift_update(date, status):
    open_shift = input('Please enter' + status + 'shift worker name, hours, tip, (shabbat hours): ')
    open_shift_details = open_shift.split()
    shabbat_hours = float(open_shift_details[3]) if len(open_shift_details) == 4 else 0
    update_worker(wb[open_shift_details[0]], date, float(open_shift_details[1]), float(open_shift_details[2]), shabbat_hours, False)

def income_update(date):
    date_list = date.split(".")
    month = date_list[1]
    year = date_list[2]
    correct_date = months[int(month)-1]
    x = input('Hello noam! Please enter daily x: ')
    credit = input('Please enter credit amount: ')
    tip = input('Please enter tip amount:')
    update_income(wb[correct_date + " 20" + year], date, float(x), float(credit), float(tip))

def update_day(wb, is_shabbat=False):
    date = input('Please enter date:')
    income_update(date)
    if is_shabbat:
        regular_shift_update(date, " Noon ")

    regular_shift_update(date, " open ")
    regular_shift_update(date, " close ")
    update_pik(date)
    wb.save("worker tracing.xlsx")


# wb = op.Workbook()
# add_worker(wb, "Michal", categories)
# wb = op.load_workbook('worker tracing.xlsx')



# add_worker(wb, "", categories)
wb = new_workbook(worker_list, 10, 21)
# update_worker(wb['Alon'], "24.8.21", 9, 323, 0, False)
# update_worker(wb['Gilad'], "1.1.21", 9.5, 300, 0, False)
# update_worker(wb['Rinat'], "25.8.21", 4, 0, 0, True)
# update_income(wb['September 2021'], "26.8.21", 4195, 3520, 489)

update_day(wb)
# income_update('1.9.21')
